# -*- coding: utf-8 -*-
"""
Created on Wed Feb 15 14:17:00 2017

@author: a550859
"""
import os
from openpyxl import Workbook
from openpyxl.compat import range
from openpyxl.cell import get_column_letter

wb = Workbook()
dest_filename = 'empty_book.xlsx'

ws1 = wb.active
ws1.title = "range names"

for row in range(1, 40):
    ws1.append(range(600))

ws2 = wb.create_sheet(title="feb17")

ws2['A4'] ="Automation demo"

print(os.getcwd())
wb.save(filename=dest_filename)


'''

ws3 = wb.create_sheet(title="feb17")
for row in range(10, 20):
    for col in range(27, 54):
        _ = ws3.cell(column=col, row=row, value="%s" % get_column_letter(col))
print(ws3['AA10'].value)

wb.save(filename = dest_filename)
'''